from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import scikit_posthocs as sp
from scipy.stats import friedmanchisquare


PROJECT_ROOT = Path(__file__).resolve().parents[3]
INPUT = PROJECT_ROOT / "ARM with TFMs.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "statistical_analysis_results"
SHEET = "Average Results"
ALPHA = 0.05

DATASETS = [
    "breast_cancer",
    "congressional_voting",
    "mushroom",
    "chess_king_rook_vs_king_pawn",
    "spambase",
    "hepatitis",
    "cervical_cancer_behavior_risk",
    "autism_screening_adolescent",
    "acute_inflammations",
    "fertility",
]

METHOD_MAP = {
    "TabICLv2": "TabICLv2",
    "TabPFN": "TabPFN",
    "TabDPT": "TabDPT",
    "Aerial+": "Aerial+",
    "FP-Growth\nmin_supp=0.5": "FP-Growth min_supp=0.5",
    "FP-Growth\nmin_supp=0.3": "FP-Growth min_supp=0.3",
    "FP-Growth\nmin_supp=0.2": "FP-Growth min_supp=0.2",
    "FP-Growth\nmin_supp=0.1": "FP-Growth min_supp=0.1",
    "XGBoost": "XGBoost",
}

METHODS = list(METHOD_MAP.values())
METRICS = [
    "num_rules",
    "avg_support",
    "avg_confidence",
    "avg_zhangs_metric",
    "avg_interestingness",
    "data_coverage",
]


def read_block(sheet, start_col: int) -> pd.DataFrame:
    headers = [sheet.cell(2, start_col + offset).value for offset in range(9)]
    rows = []
    current_algorithm = None
    for row in range(3, 58):
        values = [sheet.cell(row, start_col + offset).value for offset in range(9)]
        if values[0] is not None:
            current_algorithm = values[0]
        dataset = values[1]
        if current_algorithm in METHOD_MAP and dataset in DATASETS:
            rows.append([METHOD_MAP[current_algorithm], *values[1:]])
    return pd.DataFrame(rows, columns=["algorithm", *headers[1:]])


def main() -> None:
    workbook = openpyxl.load_workbook(INPUT, data_only=True, read_only=False)
    sheet = workbook[SHEET]
    data = pd.concat([read_block(sheet, 1), read_block(sheet, 13)], ignore_index=True)
    data = data[["algorithm", "dataset", *METRICS]].copy()

    expected = pd.MultiIndex.from_product([DATASETS, METHODS], names=["dataset", "algorithm"])
    observed = pd.MultiIndex.from_frame(data[["dataset", "algorithm"]])
    missing = expected.difference(observed)
    duplicates = data.duplicated(["dataset", "algorithm"], keep=False)
    if len(missing) or duplicates.any():
        raise ValueError(
            f"Incomplete or duplicated design. Missing={missing.tolist()}, "
            f"duplicates={data.loc[duplicates, ['dataset', 'algorithm']].to_dict('records')}"
        )
    if data[METRICS].isna().any().any():
        raise ValueError("Selected design contains missing metric values.")

    OUTPUT_DIR.mkdir(exist_ok=True)
    data.to_csv(OUTPUT_DIR / "normalized_input.csv", index=False)

    omnibus_rows = []
    rank_rows = []
    significant_pairs = []

    for metric in METRICS:
        wide = data.pivot(index="dataset", columns="algorithm", values=metric).loc[DATASETS, METHODS]
        wide.to_csv(OUTPUT_DIR / f"input_matrix_{metric}.csv")

        statistic, p_value = friedmanchisquare(*(wide[method].to_numpy() for method in METHODS))
        kendalls_w = statistic / (len(DATASETS) * (len(METHODS) - 1))
        reject = bool(p_value < ALPHA)
        omnibus_rows.append(
            {
                "metric": metric,
                "n_datasets": len(DATASETS),
                "n_methods": len(METHODS),
                "friedman_chi_square": statistic,
                "degrees_of_freedom": len(METHODS) - 1,
                "p_value": p_value,
                "alpha": ALPHA,
                "reject_equal_ranks": reject,
                "kendalls_w": kendalls_w,
            }
        )

        # Descending ranks make rank 1 the highest value. Reversing rank direction
        # does not change Friedman or Nemenyi significance results.
        ranks = wide.rank(axis=1, method="average", ascending=False)
        mean_ranks = ranks.mean(axis=0).sort_values()
        for method, mean_rank in mean_ranks.items():
            rank_rows.append({"metric": metric, "method": method, "average_rank": mean_rank})

        if reject:
            nemenyi = sp.posthoc_nemenyi_friedman(wide)
            nemenyi = nemenyi.loc[METHODS, METHODS]
            nemenyi.to_csv(OUTPUT_DIR / f"nemenyi_pvalues_{metric}.csv")
            for i, method_a in enumerate(METHODS):
                for method_b in METHODS[i + 1 :]:
                    pair_p = float(nemenyi.loc[method_a, method_b])
                    if pair_p < ALPHA:
                        significant_pairs.append(
                            {
                                "metric": metric,
                                "method_a": method_a,
                                "method_b": method_b,
                                "average_rank_a": float(mean_ranks[method_a]),
                                "average_rank_b": float(mean_ranks[method_b]),
                                "absolute_rank_difference": float(abs(mean_ranks[method_a] - mean_ranks[method_b])),
                                "nemenyi_p_value": pair_p,
                                "alpha": ALPHA,
                            }
                        )

    omnibus = pd.DataFrame(omnibus_rows)
    average_ranks = pd.DataFrame(rank_rows)
    pairs = pd.DataFrame(
        significant_pairs,
        columns=[
            "metric",
            "method_a",
            "method_b",
            "average_rank_a",
            "average_rank_b",
            "absolute_rank_difference",
            "nemenyi_p_value",
            "alpha",
        ],
    )
    omnibus.to_csv(OUTPUT_DIR / "friedman_results.csv", index=False)
    average_ranks.to_csv(OUTPUT_DIR / "average_ranks.csv", index=False)
    pairs.to_csv(OUTPUT_DIR / "nemenyi_significant_pairs.csv", index=False)

    lines = [
        "# Friedman and Nemenyi analysis",
        "",
        f"Source: `{INPUT.name}`, sheet `{SHEET}`.",
        f"Design: {len(DATASETS)} datasets (blocks) × {len(METHODS)} methods; alpha = {ALPHA} per metric.",
        "Rank 1 denotes the highest observed value (descriptive, not necessarily best).",
        "Nemenyi tests are performed only when the corresponding Friedman test rejects equal ranks.",
        "",
        "## Friedman tests",
        "",
        omnibus.to_markdown(index=False, floatfmt=".6g"),
        "",
        "## Significant Nemenyi pairs",
        "",
        pairs.to_markdown(index=False, floatfmt=".6g") if not pairs.empty else "No significant pairs.",
        "",
        "## Average ranks",
        "",
        average_ranks.pivot(index="method", columns="metric", values="average_rank")
        .loc[METHODS, METRICS]
        .to_markdown(floatfmt=".3f"),
        "",
    ]
    (OUTPUT_DIR / "analysis_report.md").write_text("\n".join(lines), encoding="utf-8")

    print(omnibus.to_string(index=False))
    print("\nSignificant Nemenyi pairs:")
    print(pairs.to_string(index=False) if not pairs.empty else "None")


if __name__ == "__main__":
    main()
